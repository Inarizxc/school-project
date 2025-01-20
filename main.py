import json
import tkinter as tk
from tkinter import filedialog

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd


def clear_canvas():
    for widget in root.winfo_children():
        if widget != button:
            widget.destroy()


def open_file():
    file_path = filedialog.askopenfilename(
        filetypes=[
            ("JSON files", "*.json"),
            ("CSV files", "*.csv"),
            ("Excel files", "*.xlsx"),
        ]
    )
    if not file_path:
        return

    try:
        if file_path.endswith(".json"):
            with open(file_path, "r") as f:
                data = json.load(f)

            key_var = tk.StringVar(value="")
            key_label = tk.Label(root, text="Ключ:")
            key_entry = tk.Entry(root, textvariable=key_var)
            key_label.pack()
            key_entry.pack()

            def extract_and_plot():
                key = key_var.get()
                if key:
                    extracted_data = data.get(key, [])
                    df = pd.DataFrame(extracted_data)
                else:
                    df = pd.DataFrame(data)

                clear_canvas()

                x_column = tk.StringVar(value="x")
                y_column = tk.StringVar(value="y")
                kind_var = tk.StringVar(value="line")
                color_var = tk.StringVar(value="blue")
                width_var = tk.DoubleVar(value=1)

                def create_graph():
                    x_col = x_column.get()
                    y_col = y_column.get()
                    kind = kind_var.get()
                    color = color_var.get()
                    width = width_var.get()

                    if x_col not in df.columns or y_col not in df.columns:
                        show_error(f"'{x_col}' и/или '{y_col}' отсутствуют в файле.")
                        return

                    plot_data(df, x_col, y_col, kind, color, width)

                label_x = tk.Label(root, text="Выберите столбец 1:")
                entry_x = tk.Entry(root, textvariable=x_column)
                label_y = tk.Label(root, text="Выберите столбец 2:")
                entry_y = tk.Entry(root, textvariable=y_column)
                radio_line = tk.Radiobutton(
                    root, variable=kind_var, value="line", text="Линейный"
                )
                radio_scatter = tk.Radiobutton(
                    root, variable=kind_var, value="scatter", text="Точечный"
                )
                color_picker = tk.Frame(root)
                color_picker.pack()
                color_label = tk.Label(color_picker, text="Цвет: ")
                color_entry = tk.Entry(color_picker, textvariable=color_var)
                width_slider = tk.Scale(
                    root, variable=width_var, from_=0.1, to=5, orient="horizontal"
                )
                button_plot = tk.Button(
                    root, text="Построить график", command=create_graph
                )

                label_x.pack()
                entry_x.pack()
                label_y.pack()
                entry_y.pack()
                radio_line.pack()
                radio_scatter.pack()
                color_label.pack()
                color_entry.pack()
                width_slider.pack()
                button_plot.pack()

                key_label.destroy()
                key_entry.destroy()

            extract_button = tk.Button(
                root, text="Получить данные", command=extract_and_plot
            )
            extract_button.pack()

        elif file_path.endswith(".csv"):
            df = pd.read_csv(file_path)

            clear_canvas()

            x_column = tk.StringVar(value="x")
            y_column = tk.StringVar(value="y")
            kind_var = tk.StringVar(value="line")
            color_var = tk.StringVar(value="blue")
            width_var = tk.DoubleVar(value=1)

            def create_graph():
                x_col = x_column.get()
                y_col = y_column.get()
                kind = kind_var.get()
                color = color_var.get()
                width = width_var.get()

                if x_col not in df.columns or y_col not in df.columns:
                    show_error(f"'{x_col}' и/или '{y_col}' отсутствуют в данных.")
                    return

                plot_data(df, x_col, y_col, kind, color, width)

            label_x = tk.Label(root, text="Выберите столбец 1:")
            entry_x = tk.Entry(root, textvariable=x_column)
            label_y = tk.Label(root, text="Выберите столбец 2:")
            entry_y = tk.Entry(root, textvariable=y_column)
            radio_line = tk.Radiobutton(
                root, variable=kind_var, value="line", text="Линейный"
            )
            radio_scatter = tk.Radiobutton(
                root, variable=kind_var, value="scatter", text="Точечный"
            )
            color_picker = tk.Frame(root)
            color_picker.pack()
            color_label = tk.Label(color_picker, text="Цвет: ")
            color_entry = tk.Entry(color_picker, textvariable=color_var)
            width_slider = tk.Scale(
                root, variable=width_var, from_=0.1, to=5, orient="horizontal"
            )
            button_plot = tk.Button(root, text="Построить график", command=create_graph)

            label_x.pack()
            entry_x.pack()
            label_y.pack()
            entry_y.pack()
            radio_line.pack()
            radio_scatter.pack()
            color_label.pack()
            color_entry.pack()
            width_slider.pack()
            button_plot.pack()

        elif file_path.endswith(".xlsx"):
            df = load_subset(pd.read_excel(file_path))

            clear_canvas()

            x_column = tk.StringVar(value="x")
            y_column = tk.StringVar(value="y")
            kind_var = tk.StringVar(value="line")
            color_var = tk.StringVar(value="blue")
            width_var = tk.DoubleVar(value=1)

            def create_graph():
                x_col = x_column.get()
                y_col = y_column.get()
                kind = kind_var.get()
                color = color_var.get()
                width = width_var.get()

                if x_col not in df.columns or y_col not in df.columns:
                    show_error(f"'{x_col}' и/или '{y_col}' отсутствуют в файле.")
                    return

                plot_data(df, x_col, y_col, kind, color, width)

            label_x = tk.Label(root, text="Выберите столбец 1:")
            entry_x = tk.Entry(root, textvariable=x_column)
            label_y = tk.Label(root, text="Выберите столбец 2:")
            entry_y = tk.Entry(root, textvariable=y_column)
            radio_line = tk.Radiobutton(
                root, variable=kind_var, value="line", text="Линейный"
            )
            radio_scatter = tk.Radiobutton(
                root, variable=kind_var, value="scatter", text="Точечный"
            )
            color_picker = tk.Frame(root)
            color_picker.pack()
            color_label = tk.Label(color_picker, text="Цвет: ")
            color_entry = tk.Entry(color_picker, textvariable=color_var)
            width_slider = tk.Scale(
                root, variable=width_var, from_=0.1, to=5, orient="horizontal"
            )
            button_plot = tk.Button(root, text="Построить график", command=create_graph)

            label_x.pack()
            entry_x.pack()
            label_y.pack()
            entry_y.pack()
            radio_line.pack()
            radio_scatter.pack()
            color_label.pack()
            color_entry.pack()
            width_slider.pack()
            button_plot.pack()

        else:
            raise ValueError("Неподдерживаемый формат файла.")
    except Exception as e:
        show_error(f"Ошибка при чтении файла: {e}")


def show_error(message):
    error_box = tk.Toplevel(root)
    error_box.title("Ошибка")
    label = tk.Label(error_box, text=message, font=("Helvetica", 12), justify="center")
    label.pack(padx=20, pady=20)
    close_button = tk.Button(error_box, text="Закрыть", command=error_box.destroy)
    close_button.pack(pady=10)


def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    data = []
    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        data.append(row_data)

    return data


def load_subset(df, n=1000):
    return df.head(n)


def plot_data(df, x_col, y_col, kind="line", color="blue", width=1):
    plt.figure(figsize=(6, 4))

    if kind == "line":
        plt.plot(df[x_col], df[y_col], color=color, linewidth=width)
    elif kind == "scatter":
        plt.scatter(df[x_col], df[y_col], color=color, s=width)

    plt.xlabel(x_col)
    plt.ylabel(y_col)
    plt.title(f"График из файла ({x_col}, {y_col}) - {kind}")
    plt.show()


root = tk.Tk()
root.title("Большие данные")

button = tk.Button(root, text="Загрузить файл", command=open_file)
button.pack(pady=10)

root.mainloop()

